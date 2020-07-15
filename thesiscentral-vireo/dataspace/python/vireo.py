import logging
import openpyxl
from openpyxl import Workbook
import pandas as pd

class VireoSheet:
    """
    An Excel Spreadsheet encoding information for student thesis submissions.

    Attributes
    ----------
    file_name : string
        The name of the Excel spreadsheet file.

    Methods
    -------
    col_names()
        Retrieve all of the names of the spreadsheet columns.
    """

    CERTIFICATE_PROGRAM = 'Certificate Program'
    STUDENT_NAME = 'Student name'
    STUDENT_ID = 'Student ID'
    DEPARTMENT = 'Department'
    SUBMISSION_DATE = 'Submission date'
    ADVISORS = 'Advisors'
    DOCUMENT_TITLE = 'Title'
    PRIMARY_DOCUMENT = 'Primary document'
    THESIS_TYPE = 'Thesis Type'
    STUDENT_EMAIL = 'Student email'
    MULTI_AUTHOR = 'Multi Author'
    STATUS = 'Status'
    TITLE = 'Title'
    ID = 'ID'

    # RESTRICTIONS SHEET
    R_STUDENT_NAME = 'Submitted By'
    R_TITLE = 'Name'
    R_WALK_IN = 'Walk In Access'
    R_EMBARGO = 'Embargo Years'

    def __init__(self, workbook, name, unique_ids=True):
        self.file_name = name
        self.unique_ids = unique_ids

        if (workbook):
            # This is only referenced for the workbook and to save the spreadsheet
            self._thesis_wb = workbook

            if (1 != len(self._thesis_wb.worksheets)) :
                raise Exception("%s should have exactly one sheet" % (self.file_name))

            # This is what is directly modified in the Excel Spreadsheet
            self._sheet = self._thesis_wb.worksheets[0]
            self.id_col = self.col_index_of(VireoSheet.ID, required=True)
            self.id_rows = self._derive_id_hash()
            logging.debug("IDS: " +  ", ".join(str(_) for _  in self.id_rows))
        else:
            raise Exception("invalid workbook")

    @staticmethod
    def createFromExcelFile(file_name, unique_ids=True):
        """
        Factory method for constructing VireoSheet objects from Excel spreadsheet workbook

        Parameters
        ----------
        file_name : string
            The file system path to the Excel spreadsheet
        unique_ids : bool
            Whether or not to enforce unique IDs in the spreadsheet rows

        Returns
        -------
        VireoSheet
            The newly-constructed VireoSheet from the Excel spreadsheet
        """
        if (file_name):
            thesis_wb = openpyxl.load_workbook(filename = file_name)
            return VireoSheet(thesis_wb, file_name, unique_ids)
        else:
            raise Exception("must give non empty filename")

    def col_names(self):
        """ 
        :return:  values from first row 
        """
        hdrs = next(self._sheet.iter_rows(min_row=1, max_row=1))
        return VireoSheet.row_values(hdrs)

    def col_index_of(self, col_header, required=False):
        """
        return index (0...n) of column headed by given name

        if required check throw expection when there is no matching column
        otherwise return None

        :param col_header:
        :param required: True or False
        :return:
        """
        try:
            return self.col_names().index(col_header)
        except ValueError as e:
            if (required):
                raise Exception("%s does not contain a '%s' column" % (self.file_name, col_header))
        return None;

    def iter_rows(self):
        return self._sheet.iter_rows(min_row=2)

    @staticmethod
    def row_values(row):
        """
        return rows value as stripped strings
        :param row: a row from the spreadsheet
        :return: list of strings
        """
        return [unicode(cell.value).strip() for cell in row]

    def matchingRows(self, col_name, value):
        """
        find rows with matching column values

        :param col_name: unique column name
        :param value: value to match on
        :return: list of rows that have value in the colum headed by col_name
        """
        matches = []
        col_idx = self.col_index_of(col_name)
        iter = self._sheet.iter_rows()
        _ = next(iter)  # throw header row away
        for row in iter:
            if (row[col_idx].value == value):
                matches.append(row)
        return matches

    def save(self, file_name):
        # Use pandas to save the data set
        dataset = {}
        columns = [
            'Name',
            'Submitted By',
            'Created',
            'Class Year',
            'Department',
            'Adviser',
            'Embargo Years',
            'Walk In Access',
            'Initial Review',
            'Adviser Comments Status',
            'Adviser Comments',
            'ODOCReview',
            'Confirmation Sent',
            'Approval Notification Sent',
            'Mudd Status',
            'Request Type',
            'Notify Faculty Adviser',
            'Thesis Uploaded',
            'Check Thesis Uploaded',
            'SetFormLink',
            'Item Type',
            'Path'
        ]

        for col in columns:
            dataset[col] = []

        row_index = 0
        for row in self._sheet.iter_rows(min_row=2):
            keys = dataset.keys()
            cell_index = 0

            for cell in row:
                try:
                    column = columns[cell_index - 1]

                    # @todo investigate why this is breaking pandas
                    if cell_index == 22:
                        continue

                    dataset[column].append(cell.value)
                    cell_index += 1
                except Exception as inst:
                    import pdb; pdb.set_trace()
                    pass

            row_index += 1

        data_frame = pd.DataFrame(dataset)
        logging.info("SAVING to: " + file_name)

        self._thesis_wb.save(file_name)

    def _derive_id_hash(self):
        id_rows = {}

        for row in self._sheet.iter_rows(min_row=2):
            try:
                cell = row[self.id_col]
                if not cell.value:
                    first_field = str(row[0].value.encode('utf-8'))
                    logging.warning("%s: Row has no ID value: %s" % (self.file_name, first_field))
                else:
                    row_id = int(row[self.id_col].value)
                    if not row_id in id_rows:
                        id_rows[row_id] = [row]
                    elif self.unique_ids:
                        raise Exception("%s has duplicate id %s" % (self.file_name, str(row_id)))
                    else:
                        id_rows[row_id].append(row)
            except Exception as inst:
                import pdb; pdb.set_trace()
                pass

        return id_rows

    def readMoreCerts(self, add_certs_file, check_id=True):
        logging.info("readMoreCerts")
        add_certs = VireoSheet.createFromExcelFile(add_certs_file, unique_ids=False)
        # check that required columns are present
        certs_email_col_id = add_certs.col_index_of(VireoSheet.STUDENT_EMAIL, required=True)
        certs_cert_col_id = add_certs.col_index_of(VireoSheet.CERTIFICATE_PROGRAM, required=True)
        thesis_email_col_id = self.col_index_of(VireoSheet.STUDENT_EMAIL, required=True)
        # look through whether certs file info matches thesis sheet info
        for cert_id in add_certs.id_rows:
            if  not  cert_id in  self.id_rows:
                if (check_id):
                    raise Exception("%s, row with ID %d: there is no such row in %s" % (add_certs.file_name, cert_id, self.file_name))
                else:
                    continue
            thesis_row = self.id_rows[cert_id][0]
            for row in  add_certs.id_rows[cert_id]:
                logging.debug("ID %d -> cert_row: %s" % (cert_id, VireoSheet.row_values(row)))
                # check that student email matches
                if (row[certs_email_col_id].value.strip() != thesis_row[thesis_email_col_id].value.strip()):
                    raise Exception("%s, row with ID %d: email mistmatch with same row in %s" % (add_certs.file_name, cert_id, self.file_name))
                # check that certificate program enry is not empty
                if not row[certs_cert_col_id].value:
                    raise Exception("%s, row with ID %d: row has empty certificate value" % (add_certs.file_name, cert_id))
        return add_certs

    def readRestrictions(self, restriction_file, check_id=True):
        """
        Parses the restrictions Excel spreadsheet provided by Department of the Registrar

        Parameters
        ----------
        restriction_file : string
            The path to the Excel restriction file

        check_id : bool
            Whether or not to check for restriction IDs which are not in the submission IDs

        Returns
        -------
        VireoSheet
            The VireoSheet object constructed from the parsed Excel Spreadsheet

        """

        restrictions = VireoSheet.createFromExcelFile(restriction_file, unique_ids=False)
        # check that necessary columns are present
        restrictions.col_index_of(VireoSheet.R_STUDENT_NAME, required=True)
        restrictions.col_index_of(VireoSheet.R_TITLE, required=True)
        restrictions.col_index_of(VireoSheet.R_WALK_IN, required=True)
        restrictions.col_index_of(VireoSheet.R_EMBARGO, required=True)
        # check whether restriction IDs make sense
        # look through whether certs file info matches thesis sheet info
        if (check_id):
            for restr_id in restrictions.id_rows:
                if  not  restr_id in  self.id_rows:
                    raise Exception("%s, row with ID %d: there is no such row in %s" % (restrictions.file_name, restr_id, self.file_name))
        return restrictions

    def log_info(self):
        logging.info(str(self))
        logging.info("%s: headers %s" % (self.file_name, str.join(',', self.col_names())))
        for col in [VireoSheet.STUDENT_EMAIL, VireoSheet.CERTIFICATE_PROGRAM]:
            logging.info("%s column: %s (%s)" % (self.file_name, col, str(self.col_index_of(col))))
        logging.info("---")

    def __str__(self):
        return "%s: ID%s-col:%d, nrows:%d" % (self.file_name, ('(unique)' if self.unique_ids else ''), self.id_col, len(self.id_rows))

def createWithAddedColumn(file_name, col_name, unique_ids=True):
    """
    read excel sheet from file and create VireoSheet with added column

    pass unique_ids parameter on to VireoSheet constructor

    :param file_name: excel file with spreadsheet containing a single worksheet
    :param col_name: name of additional column
    :return: VireoSheet
    """

    from_wb = openpyxl.load_workbook(filename = file_name)

    if (1 != len(from_wb.worksheets)) :
        raise Exception("%s should have exactly one sheet" % (file_name))

    new_name = "%s_+_%s-column"  % (file_name, col_name)
    wb = Workbook(new_name)
    wb_sheet = wb.create_sheet("MainSheet")
    for row in from_wb.worksheets[0]:
        tpl = row + (None, )
        wb_sheet.append(tpl)

    return VireoSheet(wb, new_name, unique_ids)
